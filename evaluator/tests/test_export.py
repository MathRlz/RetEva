"""Tests for result export functionality."""

import csv
import json
import tempfile
from pathlib import Path
from typing import Any, Dict

import pytest

from evaluator.analysis.export import (
    export_to_csv,
    export_to_excel,
    export_to_latex,
    compare_experiments_to_latex,
    export_sample_results,
    load_results,
    _escape_latex,
)


class TestExportToCsv:
    """Tests for export_to_csv function."""

    def test_exports_basic_metrics(self):
        """Should export numeric metrics to CSV."""
        results = {"MRR": 0.85, "Recall@5": 0.92, "WER": 0.15}
        
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False
        ) as f:
            path = f.name
        
        try:
            export_to_csv(results, path)
            
            with open(path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            
            metric_names = [row["metric"] for row in rows]
            assert "MRR" in metric_names
            assert "Recall@5" in metric_names
            assert "WER" in metric_names
        finally:
            Path(path).unlink()

    def test_exports_metadata_fields(self):
        """Should include metadata fields like asr and embedder."""
        results = {
            "asr": "WhisperModel",
            "embedder": "LaBSE",
            "MRR": 0.85
        }
        
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False
        ) as f:
            path = f.name
        
        try:
            export_to_csv(results, path)
            
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            
            assert "asr" in content
            assert "WhisperModel" in content
            assert "embedder" in content
            assert "LaBSE" in content
        finally:
            Path(path).unlink()

    def test_skips_complex_fields(self):
        """Should skip per_sample and details fields."""
        results = {
            "MRR": 0.85,
            "per_sample": {"MRR": [0.8, 0.9]},
            "details": [{"query": "q1"}]
        }
        
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False
        ) as f:
            path = f.name
        
        try:
            export_to_csv(results, path)
            
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            
            assert "per_sample" not in content
            assert "details" not in content
        finally:
            Path(path).unlink()

    def test_creates_parent_directories(self):
        """Should create parent directories if they don't exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "nested" / "dir" / "results.csv"
            
            results = {"MRR": 0.85}
            export_to_csv(results, str(output_path))
            
            assert output_path.exists()


class TestExportToExcel:
    """Tests for export_to_excel function."""

    @pytest.fixture
    def check_openpyxl(self):
        """Skip tests if openpyxl is not installed."""
        pytest.importorskip("openpyxl")

    def test_creates_summary_sheet(self, check_openpyxl):
        """Should create summary sheet with metrics."""
        from openpyxl import load_workbook
        
        results = {"MRR": 0.85, "Recall@5": 0.92}
        
        with tempfile.NamedTemporaryFile(
            mode="wb", suffix=".xlsx", delete=False
        ) as f:
            path = f.name
        
        try:
            export_to_excel(results, path)
            
            wb = load_workbook(path)
            assert "Summary" in wb.sheetnames
            
            ws = wb["Summary"]
            values = [cell.value for row in ws.iter_rows() for cell in row]
            assert "MRR" in values
            assert "Recall@5" in values
        finally:
            Path(path).unlink()

    def test_creates_per_sample_sheet(self, check_openpyxl):
        """Should create per-sample sheet when per_sample data exists."""
        from openpyxl import load_workbook
        
        results = {
            "MRR": 0.85,
            "per_sample": {"MRR": [0.8, 0.85, 0.9]}
        }
        
        with tempfile.NamedTemporaryFile(
            mode="wb", suffix=".xlsx", delete=False
        ) as f:
            path = f.name
        
        try:
            export_to_excel(results, path)
            
            wb = load_workbook(path)
            assert "Per-Sample" in wb.sheetnames
        finally:
            Path(path).unlink()

    def test_creates_per_sample_from_details(self, check_openpyxl):
        """Should create per-sample sheet from details list."""
        from openpyxl import load_workbook
        
        results = {
            "MRR": 0.85,
            "details": [
                {"query": "q1", "MRR": 0.8},
                {"query": "q2", "MRR": 0.9}
            ]
        }
        
        with tempfile.NamedTemporaryFile(
            mode="wb", suffix=".xlsx", delete=False
        ) as f:
            path = f.name
        
        try:
            export_to_excel(results, path)
            
            wb = load_workbook(path)
            assert "Per-Sample" in wb.sheetnames
        finally:
            Path(path).unlink()


class TestExportToLatex:
    """Tests for export_to_latex function."""

    def test_creates_valid_latex_table(self):
        """Should create valid booktabs-style table."""
        results = {"MRR": 0.85, "Recall@5": 0.92}
        
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".tex", delete=False
        ) as f:
            path = f.name
        
        try:
            export_to_latex(results, path)
            
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            
            assert "\\begin{table}" in content
            assert "\\end{table}" in content
            assert "\\toprule" in content
            assert "\\midrule" in content
            assert "\\bottomrule" in content
            assert "MRR" in content
        finally:
            Path(path).unlink()

    def test_includes_caption(self):
        """Should include provided caption."""
        results = {"MRR": 0.85}
        
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".tex", delete=False
        ) as f:
            path = f.name
        
        try:
            export_to_latex(results, path, caption="Test Caption")
            
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            
            assert "Test Caption" in content
        finally:
            Path(path).unlink()

    def test_formats_floats_to_4_decimals(self):
        """Should format float values to 4 decimal places."""
        results = {"MRR": 0.8512345678}
        
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".tex", delete=False
        ) as f:
            path = f.name
        
        try:
            export_to_latex(results, path)
            
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            
            assert "0.8512" in content
        finally:
            Path(path).unlink()


class TestCompareExperimentsToLatex:
    """Tests for compare_experiments_to_latex function."""

    def test_creates_comparison_table(self):
        """Should create table with multiple experiment columns."""
        results_a = {"MRR": 0.80, "Recall@5": 0.85}
        results_b = {"MRR": 0.85, "Recall@5": 0.90}
        
        latex = compare_experiments_to_latex(
            [results_a, results_b],
            ["Baseline", "Improved"],
            caption="Model Comparison"
        )
        
        assert "Baseline" in latex
        assert "Improved" in latex
        assert "MRR" in latex
        assert "Recall" in latex
        assert "0.8000" in latex
        assert "0.8500" in latex

    def test_saves_to_file_when_path_provided(self):
        """Should save to file when output_path is provided."""
        results_a = {"MRR": 0.80}
        results_b = {"MRR": 0.85}
        
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".tex", delete=False
        ) as f:
            path = f.name
        
        try:
            compare_experiments_to_latex(
                [results_a, results_b],
                ["A", "B"],
                output_path=path
            )
            
            assert Path(path).exists()
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            assert "MRR" in content
        finally:
            Path(path).unlink()

    def test_raises_on_mismatched_lengths(self):
        """Should raise ValueError when results and names lengths differ."""
        results_a = {"MRR": 0.80}
        results_b = {"MRR": 0.85}
        
        with pytest.raises(ValueError, match="same length"):
            compare_experiments_to_latex(
                [results_a, results_b],
                ["Only One Name"]
            )

    def test_raises_on_empty_results(self):
        """Should raise ValueError for empty results list."""
        with pytest.raises(ValueError, match="cannot be empty"):
            compare_experiments_to_latex([], [])

    def test_handles_missing_metrics(self):
        """Should show -- for missing metrics."""
        results_a = {"MRR": 0.80, "WER": 0.15}
        results_b = {"MRR": 0.85}  # Missing WER
        
        latex = compare_experiments_to_latex(
            [results_a, results_b],
            ["A", "B"]
        )
        
        assert "--" in latex


class TestExportSampleResults:
    """Tests for export_sample_results function."""

    def test_exports_per_sample_data(self):
        """Should export per_sample data to CSV."""
        results = {
            "MRR": 0.85,
            "per_sample": {
                "MRR": [0.8, 0.85, 0.9],
                "Recall@5": [0.9, 0.95, 1.0]
            }
        }
        
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False
        ) as f:
            path = f.name
        
        try:
            export_sample_results(results, path)
            
            with open(path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            
            assert len(rows) == 3
            assert "sample_id" in reader.fieldnames
            assert "MRR" in reader.fieldnames
        finally:
            Path(path).unlink()

    def test_exports_details_data(self):
        """Should export details list to CSV."""
        results = {
            "MRR": 0.85,
            "details": [
                {"query": "q1", "MRR": 0.8, "hit": True},
                {"query": "q2", "MRR": 0.9, "hit": True}
            ]
        }
        
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False
        ) as f:
            path = f.name
        
        try:
            export_sample_results(results, path)
            
            with open(path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            
            assert len(rows) == 2
            assert "query" in reader.fieldnames
        finally:
            Path(path).unlink()

    def test_raises_when_no_per_sample_data(self):
        """Should raise ValueError when no per-sample data exists."""
        results = {"MRR": 0.85}
        
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False
        ) as f:
            path = f.name
        
        try:
            with pytest.raises(ValueError, match="No per-sample data"):
                export_sample_results(results, path)
        finally:
            if Path(path).exists():
                Path(path).unlink()


class TestLoadResults:
    """Tests for load_results function."""

    def test_loads_valid_json(self):
        """Should load valid JSON file."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False
        ) as f:
            json.dump({"MRR": 0.85}, f)
            path = f.name
        
        try:
            results = load_results(path)
            assert results == {"MRR": 0.85}
        finally:
            Path(path).unlink()

    def test_raises_on_missing_file(self):
        """Should raise FileNotFoundError for missing file."""
        with pytest.raises(FileNotFoundError):
            load_results("/nonexistent/path.json")

    def test_raises_on_invalid_json(self):
        """Should raise JSONDecodeError for invalid JSON."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False
        ) as f:
            f.write("not valid json")
            path = f.name
        
        try:
            with pytest.raises(json.JSONDecodeError):
                load_results(path)
        finally:
            Path(path).unlink()


class TestEscapeLatex:
    """Tests for _escape_latex helper function."""

    def test_escapes_underscores(self):
        """Should escape underscores."""
        assert _escape_latex("Recall_at_5") == "Recall\\_at\\_5"

    def test_escapes_percent(self):
        """Should escape percent signs."""
        assert _escape_latex("100%") == "100\\%"

    def test_escapes_ampersand(self):
        """Should escape ampersands."""
        assert _escape_latex("A & B") == "A \\& B"

    def test_escapes_hash(self):
        """Should escape hash signs."""
        assert _escape_latex("#1") == "\\#1"

    def test_escapes_at_sign(self):
        """Should escape at signs."""
        assert _escape_latex("Recall@5") == "Recall{@}5"


class TestExportCLI:
    """Tests for export CLI functionality."""

    def test_cli_csv_export(self):
        """Test CLI CSV export via main function."""
        from evaluator.cli.export import main
        
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "results.json"
            output_path = Path(tmpdir) / "results.csv"
            
            input_path.write_text(json.dumps({
                "MRR": 0.85,
                "Recall@5": 0.92
            }))
            
            exit_code = main([
                str(input_path),
                "--format", "csv",
                "--output", str(output_path)
            ])
            
            assert exit_code == 0
            assert output_path.exists()

    def test_cli_latex_compare(self):
        """Test CLI LaTeX comparison export."""
        from evaluator.cli.export import main
        
        with tempfile.TemporaryDirectory() as tmpdir:
            input_a = Path(tmpdir) / "a.json"
            input_b = Path(tmpdir) / "b.json"
            output_path = Path(tmpdir) / "comparison.tex"
            
            input_a.write_text(json.dumps({"MRR": 0.80}))
            input_b.write_text(json.dumps({"MRR": 0.85}))
            
            exit_code = main([
                str(input_a), str(input_b),
                "--format", "latex-compare",
                "--output", str(output_path),
                "--names", "Baseline", "Improved"
            ])
            
            assert exit_code == 0
            assert output_path.exists()
            
            content = output_path.read_text()
            assert "Baseline" in content
            assert "Improved" in content

    def test_cli_missing_input_file(self):
        """Test CLI with missing input file."""
        from evaluator.cli.export import main
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "results.csv"
            
            exit_code = main([
                "/nonexistent/file.json",
                "--format", "csv",
                "--output", str(output_path)
            ])
            
            assert exit_code == 1

    def test_cli_latex_compare_requires_multiple_files(self):
        """Test that latex-compare requires multiple files."""
        from evaluator.cli.export import main
        
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "results.json"
            output_path = Path(tmpdir) / "comparison.tex"
            
            input_path.write_text(json.dumps({"MRR": 0.85}))
            
            exit_code = main([
                str(input_path),
                "--format", "latex-compare",
                "--output", str(output_path)
            ])
            
            assert exit_code == 1

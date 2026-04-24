"""Export evaluation results to various formats.

Provides functions for exporting results to CSV, Excel, and LaTeX formats.
"""

import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union


def export_to_csv(results: Dict[str, Any], output_path: str) -> None:
    """Export metrics to CSV format.
    
    Creates a flat CSV with one row per metric.
    
    Args:
        results: Results dictionary with metric names as keys.
        output_path: Path to output CSV file.
        
    Example output:
        metric,value
        MRR,0.85
        Recall@5,0.92
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    # Separate metadata from metrics
    skip_keys = {"asr", "embedder", "per_sample", "details", "config", 
                 "metadata", "pipeline_mode", "phased", "audio_embedder", 
                 "text_embedder"}
    
    rows = []
    
    # Add metadata rows first
    for key in ["asr", "embedder", "audio_embedder", "text_embedder", 
                "pipeline_mode"]:
        if key in results:
            rows.append({"metric": key, "value": str(results[key])})
    
    # Add numeric metrics
    for key, value in results.items():
        if key in skip_keys:
            continue
        if isinstance(value, (int, float)):
            rows.append({"metric": key, "value": value})
    
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["metric", "value"])
        writer.writeheader()
        writer.writerows(rows)


def export_to_excel(results: Dict[str, Any], output_path: str) -> None:
    """Export results to Excel with multiple sheets.
    
    Creates an Excel file with:
    - Summary sheet: Aggregate metrics
    - Per-sample sheet: Individual query results (if available)
    
    Args:
        results: Results dictionary with metric names as keys.
        output_path: Path to output Excel file (.xlsx).
        
    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )
    
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", 
                               fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    ws_summary["A1"] = "Metric"
    ws_summary["B1"] = "Value"
    ws_summary["A1"].font = header_font
    ws_summary["B1"].font = header_font
    ws_summary["A1"].fill = header_fill
    ws_summary["B1"].fill = header_fill
    
    skip_keys = {"per_sample", "details", "config", "metadata"}
    row = 2
    
    # Metadata section
    metadata_keys = ["asr", "embedder", "audio_embedder", "text_embedder", 
                     "pipeline_mode", "phased"]
    for key in metadata_keys:
        if key in results:
            ws_summary[f"A{row}"] = key
            ws_summary[f"B{row}"] = str(results[key])
            row += 1
    
    # Add blank row separator if we had metadata
    if row > 2:
        row += 1
    
    # Metrics section
    for key, value in results.items():
        if key in skip_keys or key in metadata_keys:
            continue
        if isinstance(value, (int, float)):
            ws_summary[f"A{row}"] = key
            ws_summary[f"B{row}"] = round(value, 6) if isinstance(value, float) else value
            row += 1
    
    # Auto-adjust column widths
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 25
    
    # Per-sample sheet (if available)
    per_sample = results.get("per_sample") or {}
    details = results.get("details") or []
    
    if per_sample:
        ws_samples = wb.create_sheet("Per-Sample")
        
        # Get all metric names
        metrics = list(per_sample.keys())
        if not metrics:
            pass
        else:
            # Headers
            ws_samples["A1"] = "Sample"
            ws_samples["A1"].font = header_font
            ws_samples["A1"].fill = header_fill
            
            for col_idx, metric in enumerate(metrics, start=2):
                cell = ws_samples.cell(row=1, column=col_idx, value=metric)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data rows
            n_samples = max(len(v) for v in per_sample.values())
            for sample_idx in range(n_samples):
                ws_samples.cell(row=sample_idx + 2, column=1, value=sample_idx + 1)
                for col_idx, metric in enumerate(metrics, start=2):
                    scores = per_sample.get(metric, [])
                    if sample_idx < len(scores):
                        value = scores[sample_idx]
                        if isinstance(value, float):
                            value = round(value, 6)
                        ws_samples.cell(row=sample_idx + 2, column=col_idx, value=value)
    
    elif details and isinstance(details, list) and details:
        ws_samples = wb.create_sheet("Per-Sample")
        
        # Get columns from first item
        if isinstance(details[0], dict):
            columns = list(details[0].keys())
            
            # Headers
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws_samples.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data rows
            for row_idx, item in enumerate(details, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = item.get(col_name)
                    if isinstance(value, float):
                        value = round(value, 6)
                    ws_samples.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(path)


def export_to_latex(
    results: Dict[str, Any],
    output_path: str,
    caption: str = "Evaluation Results"
) -> None:
    """Generate LaTeX table from results.
    
    Creates a booktabs-style table with proper formatting.
    
    Args:
        results: Results dictionary with metric names as keys.
        output_path: Path to output .tex file.
        caption: Table caption.
        
    Example output:
        \\begin{table}[htbp]
        \\centering
        \\caption{Evaluation Results}
        \\begin{tabular}{lr}
        \\toprule
        Metric & Value \\\\
        \\midrule
        MRR & 0.8500 \\\\
        Recall@5 & 0.9200 \\\\
        \\bottomrule
        \\end{tabular}
        \\end{table}
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    skip_keys = {"asr", "embedder", "per_sample", "details", "config", 
                 "metadata", "pipeline_mode", "phased", "audio_embedder", 
                 "text_embedder"}
    
    # Collect metrics
    metrics = []
    for key, value in results.items():
        if key in skip_keys:
            continue
        if isinstance(value, (int, float)):
            metrics.append((key, value))
    
    # Build LaTeX table
    lines = []
    lines.append("\\begin{table}[htbp]")
    lines.append("\\centering")
    lines.append(f"\\caption{{{_escape_latex(caption)}}}")
    lines.append("\\begin{tabular}{lr}")
    lines.append("\\toprule")
    lines.append("Metric & Value \\\\")
    lines.append("\\midrule")
    
    for metric_name, value in metrics:
        escaped_name = _escape_latex(metric_name)
        if isinstance(value, float):
            formatted_value = f"{value:.4f}"
        else:
            formatted_value = str(value)
        lines.append(f"{escaped_name} & {formatted_value} \\\\")
    
    lines.append("\\bottomrule")
    lines.append("\\end{tabular}")
    lines.append("\\end{table}")
    
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def compare_experiments_to_latex(
    results_list: List[Dict[str, Any]],
    names: List[str],
    output_path: Optional[str] = None,
    caption: str = "Experiment Comparison"
) -> str:
    """Generate LaTeX comparison table for multiple experiments.
    
    Creates a table with experiments as columns and metrics as rows.
    
    Args:
        results_list: List of result dictionaries.
        names: List of experiment names (same order as results_list).
        output_path: Optional path to save the .tex file.
        caption: Table caption.
        
    Returns:
        LaTeX table string.
        
    Raises:
        ValueError: If results_list and names have different lengths.
    """
    if len(results_list) != len(names):
        raise ValueError(
            f"results_list and names must have same length: "
            f"{len(results_list)} vs {len(names)}"
        )
    
    if not results_list:
        raise ValueError("results_list cannot be empty")
    
    skip_keys = {"asr", "embedder", "per_sample", "details", "config", 
                 "metadata", "pipeline_mode", "phased", "audio_embedder", 
                 "text_embedder"}
    
    # Find common numeric metrics
    all_metrics = set()
    for results in results_list:
        for key, value in results.items():
            if key not in skip_keys and isinstance(value, (int, float)):
                all_metrics.add(key)
    
    metrics = sorted(all_metrics)
    
    # Build column spec: l for metric name, then r for each experiment
    n_experiments = len(results_list)
    col_spec = "l" + "r" * n_experiments
    
    # Build header row
    escaped_names = [_escape_latex(name) for name in names]
    header = "Metric & " + " & ".join(escaped_names) + " \\\\"
    
    # Build LaTeX table
    lines = []
    lines.append("\\begin{table}[htbp]")
    lines.append("\\centering")
    lines.append(f"\\caption{{{_escape_latex(caption)}}}")
    lines.append(f"\\begin{{tabular}}{{{col_spec}}}")
    lines.append("\\toprule")
    lines.append(header)
    lines.append("\\midrule")
    
    for metric in metrics:
        escaped_metric = _escape_latex(metric)
        values = []
        for results in results_list:
            value = results.get(metric)
            if value is None:
                values.append("--")
            elif isinstance(value, float):
                values.append(f"{value:.4f}")
            else:
                values.append(str(value))
        
        row = f"{escaped_metric} & " + " & ".join(values) + " \\\\"
        lines.append(row)
    
    lines.append("\\bottomrule")
    lines.append("\\end{tabular}")
    lines.append("\\end{table}")
    
    latex_content = "\n".join(lines)
    
    if output_path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write(latex_content)
    
    return latex_content


def export_sample_results(results: Dict[str, Any], output_path: str) -> None:
    """Export per-sample results to CSV.
    
    Creates a CSV with one row per sample, including all available metrics.
    
    Args:
        results: Results dictionary containing 'per_sample' or 'details'.
        output_path: Path to output CSV file.
        
    Raises:
        ValueError: If no per-sample data is available.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    per_sample = results.get("per_sample")
    details = results.get("details")
    
    if per_sample and isinstance(per_sample, dict):
        # Format: {"metric": [scores...]}
        metrics = list(per_sample.keys())
        if not metrics:
            raise ValueError("per_sample dictionary is empty")
        
        n_samples = max(len(v) for v in per_sample.values())
        
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["sample_id"] + metrics)
            
            for i in range(n_samples):
                row = [i + 1]
                for metric in metrics:
                    scores = per_sample.get(metric, [])
                    if i < len(scores):
                        row.append(scores[i])
                    else:
                        row.append("")
                writer.writerow(row)
    
    elif details and isinstance(details, list):
        # Format: [{"metric1": val1, "metric2": val2, ...}, ...]
        if not details or not isinstance(details[0], dict):
            raise ValueError("details must be a list of dictionaries")
        
        # Get all column names from all items
        all_keys = set()
        for item in details:
            all_keys.update(item.keys())
        columns = sorted(all_keys)
        
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(details)
    
    else:
        raise ValueError(
            "No per-sample data found. Results must contain 'per_sample' "
            "or 'details' key with sample-level metrics."
        )


def load_results(path: Union[str, Path]) -> Dict[str, Any]:
    """Load evaluation results from JSON file.
    
    Args:
        path: Path to JSON results file.
        
    Returns:
        Results dictionary.
        
    Raises:
        FileNotFoundError: If file doesn't exist.
        json.JSONDecodeError: If file is not valid JSON.
    """
    path = Path(path)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _escape_latex(text: str) -> str:
    """Escape special LaTeX characters.
    
    Args:
        text: Input text.
        
    Returns:
        Text with LaTeX special characters escaped.
    """
    replacements = [
        ("\\", "\\textbackslash{}"),
        ("&", "\\&"),
        ("%", "\\%"),
        ("$", "\\$"),
        ("#", "\\#"),
        ("_", "\\_"),
        ("{", "\\{"),
        ("}", "\\}"),
        ("~", "\\textasciitilde{}"),
        ("^", "\\textasciicircum{}"),
        ("@", "{@}"),
    ]
    
    result = text
    for old, new in replacements:
        result = result.replace(old, new)
    
    return result
